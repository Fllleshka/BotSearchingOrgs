# Импорт данных из фаила dates.py
from dates import *

# Импорт библиотек
import requests
import openpyxl
from openpyxl.styles import Alignment
import os.path
import pprint

# Функция формирования общего списка данных
def colletdates():
    # Количество результатов
    results = 500
    # Количество пропущенных результатов
    skip = 0
    # Массив результатов
    massresults = []
    

    # Массив для наполения результурующего массива
    while results == 500:
        # Кладём результат запроса в переменную
        resultrequest = importdates(results, skip)
        # Пробегаемся по массиву и добавляем в результурующий массив
        for element in range(0, len(resultrequest)):
            massresults.append(resultrequest[element])
        # Выбираем сколько записей пропускаем
        skip += results
        # Записываем количество записей 
        results = len(resultrequest)
        # Ограничение из документации Яндекса
        if skip == 1500:
            break
    
    return massresults

# Функция импорта данных
def importdates(results, skip):
    # Формирование строки для запроса
    requestfull = url + "?text=" + requeststext + "&lang=ru_RU&apikey=" + apikey + "&results=" + str(results) +  "&skip=" + str(skip)
    # Формирование запроса типа GET и преобразование в формат JSON
    result = requests.get(requestfull).json()
    # Результирующий массив
    massresult = []
    # Пробегаемся по массиву и добавляем в результурующий массив нужные данные
    for element in range(0, len(result["features"])):
        massresult.append(result['features'][element]['properties'])
    # Возвращаем результат
    return massresult

# Функция создания excel фаила
def createexcelfile():
    # Если фаил уже есть, то необходимо его удалить
    if os.path.exists('export.xlsx'):
        os.remove('export.xlsx')
    # Создаём рабочую кинигу
    workbook = openpyxl.Workbook()
    # Создаём рабочий лист с названием текста запроса
    worksheet = workbook.create_sheet(requeststext, 0)
    # Сохраняем фаил как 'export.xlsx'
    workbook.save('export.xlsx')
    # Сохраняем абсолютный путь к файлу и возвращаем его
    pathres = os.path.abspath('export.xlsx')

    return pathres

def insertdates(pathfile, dates):
    # Открываем фаил для редактирования
    workbook = openpyxl.load_workbook(pathfile)
    # Открываем лист для редактирования
    worksheet = workbook.worksheets[0]

    # Данные для вставки в столбцы
    massnames = ['Id организации', 'Название организации', 'Сайт', 'Адрес', 'Телефон 1', 'Телефон 2']
    masscolumns = ['A', 'B', 'C', 'D', 'E', 'F']
    masswidth = [16, 54, 61, 103, 17, 17]

    # Вставка названий столбцов
    for element in range(0, len(massnames)):
        cellname = masscolumns[element] + str(1)
        worksheet[cellname] = massnames[element]
        worksheet.column_dimensions[masscolumns[element]].width = masswidth[element]
        worksheet[cellname].alignment = Alignment(horizontal='center', vertical='center')
    # Настойка высоты первой стоки
    worksheet.row_dimensions[1].height = 35
    # Фиксируем все, что левее и выше ячейки "G2"
    worksheet.freeze_panes = "G2"

    print(f"Количество импортированных записей: {len(dates)}")
    resultrequest = []
    for element in dates:
        data = element["CompanyMetaData"]
        id = data["id"]
        name = data["name"]
        try:
            urlorg = data['url']
        except:
            urlorg = "нету"

        address = data["address"]
        try:
            phone1 = data["Phones"][0]['formatted']
            try:
                phone2 = data["Phones"][1]['formatted']
            except:
                phone2 = "Отсутствует"      
        except:
            phone1 = "Отсутствует"
            phone2 = "Отсутствует"

        worksheet.append({1:id, 2:name, 3:urlorg, 4:address, 5:phone1, 6:phone2})

    # Сохраняем фаил как 'export.xlsx'
    workbook.save(pathfile)